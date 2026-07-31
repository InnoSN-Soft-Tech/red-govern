"""Local Excel governance reporting for Red-Govern."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from red_govern import __version__
from red_govern.analyzers import ObjectQuotaAnalysis
from red_govern.capabilities import CapabilityReport
from red_govern.classification import ClassificationResult
from red_govern.collectors import ObjectInventoryResult
from red_govern.exceptions import ReportError

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

SECTION_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)

ERROR_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC",
)

SUCCESS_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAD3",
)


def _style_header_row(
    worksheet: Any,
    row_number: int = 1,
) -> None:
    """Apply standard Red-Govern header styling."""
    for cell in worksheet[row_number]:
        if cell.value is None:
            continue

        cell.fill = HEADER_FILL
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


def _auto_size_columns(
    worksheet: Any,
    *,
    maximum_width: int = 45,
) -> None:
    """Apply bounded column widths based on visible values."""
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(
            column_cells[0].column
        )

        longest = max(
            (
                len(str(cell.value))
                for cell in column_cells
                if cell.value is not None
            ),
            default=0,
        )

        worksheet.column_dimensions[column_letter].width = min(
            max(longest + 2, 10),
            maximum_width,
        )


def _add_table(
    worksheet: Any,
    *,
    name: str,
) -> None:
    """Convert the populated worksheet area into an Excel table."""
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return

    reference = (
        f"A1:"
        f"{get_column_letter(worksheet.max_column)}"
        f"{worksheet.max_row}"
    )

    table = Table(
        displayName=name,
        ref=reference,
    )

    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    worksheet.add_table(table)


def _classification_lookup(
    classification: ClassificationResult | None,
) -> dict[
    tuple[str, str, str, str],
    dict[str, Any],
]:
    """Create an inventory-identity classification lookup."""
    if classification is None:
        return {}

    result: dict[
        tuple[str, str, str, str],
        dict[str, Any],
    ] = {}

    for item in classification.objects:
        identity = (
            item.record.database_name,
            item.record.schema_name,
            item.record.object_name,
            item.record.object_type.value,
        )

        result[identity] = {
            "unclassified": item.unclassified,
            "has_conflict": item.has_conflict,
            "labels": "; ".join(
                (
                    f"{dimension.dimension}="
                    f"{dimension.label}"
                )
                for dimension in item.dimensions
                if dimension.label is not None
            ),
            "conflicts": "; ".join(
                (
                    f"{dimension.dimension}: "
                    f"{', '.join(dimension.competing_rules)}"
                )
                for dimension in item.dimensions
                if dimension.conflict
            ),
        }

    return result


def _build_executive_summary(
    workbook: Workbook,
    *,
    capabilities: CapabilityReport,
    inventory: ObjectInventoryResult,
    quota: ObjectQuotaAnalysis,
    classification: ClassificationResult | None,
) -> None:
    """Create the workbook executive-summary sheet."""
    worksheet: Worksheet = workbook.create_sheet(title="Executive Summary",
        index=0,)

    worksheet["A1"] = "Red-Govern Governance Report"
    worksheet["A1"].font = Font(
        bold=True,
        size=18,
        color="FFFFFF",
    )
    worksheet["A1"].fill = HEADER_FILL
    worksheet.merge_cells("A1:D1")
    worksheet["A1"].alignment = Alignment(
        horizontal="center",
    )

    generated_at = datetime.now(timezone.utc)

    rows = [
        ("Generated at", generated_at.isoformat()),
        ("Red-Govern version", __version__),
        (
            "Deployment type",
            capabilities.deployment_type.value,
        ),
        ("Server version", capabilities.server_version),
        ("Inventory query", inventory.resolution.query.query_id),
        (
            "Inventory source",
            inventory.resolution.selected_family.value,
        ),
        ("Total objects", inventory.total_objects),
        ("Quota status", quota.status.value),
        (
            "Quota limit",
            quota.quota_limit
            if quota.quota_limit is not None
            else "Not configured",
        ),
        (
            "Remaining capacity",
            quota.remaining_capacity
            if quota.remaining_capacity is not None
            else "Unknown",
        ),
        (
            "Utilisation",
            (
                f"{quota.utilisation_percentage:.2f}%"
                if quota.utilisation_percentage is not None
                else "Unknown"
            ),
        ),
        (
            "Classified objects",
            (
                classification.classified_count
                if classification is not None
                else "Classification disabled"
            ),
        ),
        (
            "Unclassified objects",
            (
                classification.unclassified_count
                if classification is not None
                else "Classification disabled"
            ),
        ),
        (
            "Classification conflicts",
            (
                classification.conflict_count
                if classification is not None
                else "Classification disabled"
            ),
        ),
        ("Telemetry", "Disabled"),
        ("External transmission", "None"),
        ("Credentials included", "No"),
        ("Query text included", "No"),
    ]

    worksheet.append([])
    worksheet.append(["Metric", "Value"])

    for metric, value in rows:
        worksheet.append([metric, value])

    _style_header_row(worksheet, row_number=3)

    quota_status_cell = worksheet["B10"]

    if quota.status.value == "healthy":
        quota_status_cell.fill = SUCCESS_FILL
    elif quota.status.value in {"warning", "unknown"}:
        quota_status_cell.fill = WARNING_FILL
    else:
        quota_status_cell.fill = ERROR_FILL

    worksheet.freeze_panes = "A4"
    _auto_size_columns(worksheet)


def _build_inventory_sheet(
    workbook: Workbook,
    *,
    inventory: ObjectInventoryResult,
    classification: ClassificationResult | None,
) -> None:
    """Create the normalised object-inventory worksheet."""
    worksheet: Worksheet = workbook.create_sheet("Object Inventory")

    headers = [
        "Database",
        "Schema",
        "Object",
        "Object Type",
        "Owner",
        "Size MB",
        "Distribution Style",
        "Sort Key",
        "Source Family",
        "Source Query",
        "Collected At",
        "Classifications",
        "Unclassified",
        "Conflict",
    ]

    worksheet.append(headers)

    classification_lookup = _classification_lookup(
        classification
    )

    for record in inventory.records:
        identity = (
            record.database_name,
            record.schema_name,
            record.object_name,
            record.object_type.value,
        )

        classification_details = classification_lookup.get(
            identity,
            {},
        )

        worksheet.append(
            [
                record.database_name,
                record.schema_name,
                record.object_name,
                record.object_type.value,
                record.owner_name,
                record.size_mb,
                record.distribution_style,
                record.sort_key,
                record.source_family.value,
                record.source_query_id,
                record.collected_at.isoformat(),
                classification_details.get("labels"),
                classification_details.get("unclassified"),
                classification_details.get("has_conflict"),
            ]
        )

    _style_header_row(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _auto_size_columns(worksheet)
    _add_table(
        worksheet,
        name="ObjectInventoryTable",
    )


def _build_breakdown_sheet(
    workbook: Workbook,
    *,
    title: str,
    first_column: str,
    items: tuple[Any, ...],
    table_name: str,
) -> None:
    """Create a reusable quota-breakdown worksheet."""
    worksheet: Worksheet = workbook.create_sheet(title)

    worksheet.append(
        [
            first_column,
            "Object Count",
            "Percentage of Inventory",
        ]
    )

    for item in items:
        worksheet.append(
            [
                item.name,
                item.count,
                item.percentage_of_inventory / 100,
            ]
        )

    for cell in worksheet["C"][1:]:
        cell.number_format = "0.00%"

    _style_header_row(worksheet)
    worksheet.freeze_panes = "A2"
    _auto_size_columns(worksheet)
    _add_table(
        worksheet,
        name=table_name,
    )


def _build_classification_sheets(
    workbook: Workbook,
    *,
    classification: ClassificationResult | None,
) -> None:
    """Create classification, unclassified, and conflict sheets."""
    classification_sheet = workbook.create_sheet(
        "Classification"
    )

    classification_sheet.append(
        [
            "Database",
            "Schema",
            "Object",
            "Object Type",
            "Dimension",
            "Label",
            "Matched Rule",
            "Priority",
            "Conflict",
            "Competing Rules",
        ]
    )

    unclassified_sheet = workbook.create_sheet(
        "Unclassified Objects"
    )
    unclassified_sheet.append(
        [
            "Database",
            "Schema",
            "Object",
            "Object Type",
        ]
    )

    conflict_sheet = workbook.create_sheet(
        "Classification Conflicts"
    )
    conflict_sheet.append(
        [
            "Database",
            "Schema",
            "Object",
            "Dimension",
            "Selected Label",
            "Competing Rules",
        ]
    )

    if classification is not None:
        for item in classification.objects:
            for dimension in item.dimensions:
                classification_sheet.append(
                    [
                        item.record.database_name,
                        item.record.schema_name,
                        item.record.object_name,
                        item.record.object_type.value,
                        dimension.dimension,
                        dimension.label,
                        dimension.matched_rule,
                        dimension.priority,
                        dimension.conflict,
                        ", ".join(
                            dimension.competing_rules
                        ),
                    ]
                )

                if dimension.conflict:
                    conflict_sheet.append(
                        [
                            item.record.database_name,
                            item.record.schema_name,
                            item.record.object_name,
                            dimension.dimension,
                            dimension.label,
                            ", ".join(
                                dimension.competing_rules
                            ),
                        ]
                    )

            if item.unclassified:
                unclassified_sheet.append(
                    [
                        item.record.database_name,
                        item.record.schema_name,
                        item.record.object_name,
                        item.record.object_type.value,
                    ]
                )

    for worksheet, table_name in (
        (
            classification_sheet,
            "ClassificationTable",
        ),
        (
            unclassified_sheet,
            "UnclassifiedObjectsTable",
        ),
        (
            conflict_sheet,
            "ClassificationConflictsTable",
        ),
    ):
        _style_header_row(worksheet)
        worksheet.freeze_panes = "A2"
        _auto_size_columns(worksheet)
        _add_table(
            worksheet,
            name=table_name,
        )


def _build_capabilities_sheet(
    workbook: Workbook,
    *,
    capabilities: CapabilityReport,
) -> None:
    """Create the Redshift-capabilities worksheet."""
    worksheet: Worksheet = workbook.create_sheet("Capabilities")

    worksheet.append(
        [
            "Relation",
            "Family",
            "Exists",
            "Accessible",
            "Error",
        ]
    )

    for view in capabilities.views:
        worksheet.append(
            [
                view.relation,
                view.family.value,
                view.available,
                view.accessible,
                view.error,
            ]
        )

    _style_header_row(worksheet)
    worksheet.freeze_panes = "A2"
    _auto_size_columns(worksheet)
    _add_table(
        worksheet,
        name="CapabilitiesTable",
    )


def _build_privacy_sheet(
    workbook: Workbook,
) -> None:
    """Create the privacy declaration worksheet."""
    worksheet: Worksheet = workbook.create_sheet("Privacy")

    worksheet.append(["Control", "Effective Value"])
    worksheet.append(["Local-first operation", True])
    worksheet.append(["Telemetry", False])
    worksheet.append(["External transmission", False])
    worksheet.append(["Credentials included", False])
    worksheet.append(["Query text included", False])
    worksheet.append(["Customer rows scanned", False])

    _style_header_row(worksheet)
    _auto_size_columns(worksheet)
    _add_table(
        worksheet,
        name="PrivacyTable",
    )


def build_excel_workbook(
    *,
    capabilities: CapabilityReport,
    inventory: ObjectInventoryResult,
    quota: ObjectQuotaAnalysis,
    classification: ClassificationResult | None = None,
) -> Workbook:
    """Build a complete local governance workbook."""
    workbook = Workbook()

    default_worksheet = workbook.active

    if default_worksheet is not None:
        workbook.remove(default_worksheet)

    _build_executive_summary(
        workbook,
        capabilities=capabilities,
        inventory=inventory,
        quota=quota,
        classification=classification,
    )

    _build_inventory_sheet(
        workbook,
        inventory=inventory,
        classification=classification,
    )

    _build_breakdown_sheet(
        workbook,
        title="Schema Summary",
        first_column="Schema",
        items=quota.by_schema,
        table_name="SchemaSummaryTable",
    )

    _build_breakdown_sheet(
        workbook,
        title="Object Type Summary",
        first_column="Object Type",
        items=quota.by_object_type,
        table_name="ObjectTypeSummaryTable",
    )

    _build_classification_sheets(
        workbook,
        classification=classification,
    )

    _build_capabilities_sheet(
        workbook,
        capabilities=capabilities,
    )

    _build_privacy_sheet(workbook)

    return workbook


def write_excel_report(
    workbook: Workbook,
    destination: Path,
    *,
    overwrite: bool = False,
) -> Path:
    """Write a governance workbook to a local Excel file."""
    output_path = destination.expanduser().resolve()

    if output_path.exists() and not overwrite:
        raise ReportError(
            f"Report already exists: {output_path}. "
            "Use --force to replace it."
        )

    try:
        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )
        workbook.save(output_path)
    except OSError as exc:
        raise ReportError(
            f"Unable to write Excel report: {output_path}"
        ) from exc

    return output_path
