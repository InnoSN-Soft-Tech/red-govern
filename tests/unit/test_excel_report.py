"""Tests for Red-Govern Excel governance reports."""

from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from red_govern.analyzers import analyse_object_quota
from red_govern.capabilities import (
    CapabilityReport,
    DeploymentType,
    ViewFamily,
)
from red_govern.collectors import (
    DatabaseObjectType,
    ObjectInventoryRecord,
    ObjectInventoryResult,
)
from red_govern.config.models import ObjectQuotaConfig
from red_govern.exceptions import ReportError
from red_govern.query_registry import (
    QueryDefinition,
    QueryPurpose,
    QueryResolution,
)
from red_govern.reports import (
    build_excel_workbook,
    write_excel_report,
)


def build_inventory(
    *,
    object_name: str = "orders",
) -> ObjectInventoryResult:
    """Build a synthetic inventory for Excel testing."""
    collected_at = datetime.now(timezone.utc)

    query = QueryDefinition(
        query_id="object_inventory_svv_v1",
        purpose=QueryPurpose.OBJECT_INVENTORY,
        query_version="1.0.0",
        result_schema="object_inventory_v1",
        sql="SELECT 1",
        family=ViewFamily.SVV,
        deployment_types=(),
        required_relations=(),
    )

    resolution = QueryResolution(
        query=query,
        selected_family=ViewFamily.SVV,
        used_fallback=True,
        reason="Synthetic resolution.",
    )

    record = ObjectInventoryRecord(
        database_name="analytics",
        schema_name="sales",
        object_name=object_name,
        object_type=DatabaseObjectType.TABLE,
        source_family=ViewFamily.SVV,
        source_query_id=query.query_id,
        collected_at=collected_at,
    )

    return ObjectInventoryResult(
        records=(record,),
        resolution=resolution,
        collected_at=collected_at,
    )


def build_capabilities(
    *,
    server_version: str = "Redshift test",
) -> CapabilityReport:
    """Build synthetic Redshift capabilities."""
    return CapabilityReport(
        deployment_type=DeploymentType.PROVISIONED,
        server_version=server_version,
        views=(),
    )


def test_excel_workbook_contains_required_sheets() -> None:
    """Workbook should contain all mandatory governance sheets."""
    inventory = build_inventory()

    quota = analyse_object_quota(
        inventory,
        ObjectQuotaConfig(limit_override=100),
    )

    workbook = build_excel_workbook(
        capabilities=build_capabilities(),
        inventory=inventory,
        quota=quota,
    )

    assert workbook.sheetnames == [
        "Executive Summary",
        "Object Inventory",
        "Schema Summary",
        "Object Type Summary",
        "Classification",
        "Unclassified Objects",
        "Classification Conflicts",
        "Capabilities",
        "Privacy",
    ]


def test_excel_report_is_written_and_reopens(
    tmp_path: Path,
) -> None:
    """Generated workbook should be readable after saving."""
    inventory = build_inventory()

    quota = analyse_object_quota(
        inventory,
        ObjectQuotaConfig(limit_override=100),
    )

    workbook = build_excel_workbook(
        capabilities=build_capabilities(),
        inventory=inventory,
        quota=quota,
    )

    destination = tmp_path / "reports" / "governance.xlsx"

    result = write_excel_report(
        workbook,
        destination,
    )

    reopened = load_workbook(
        result,
        read_only=True,
        data_only=True,
    )

    try:
        assert "Executive Summary" in reopened.sheetnames
        assert "Object Inventory" in reopened.sheetnames

        inventory_sheet = reopened["Object Inventory"]

        assert inventory_sheet["A2"].value == "analytics"
        assert inventory_sheet["C2"].value == "orders"
    finally:
        reopened.close()


def test_excel_report_contains_no_credential_columns() -> None:
    """Workbook headers must not expose credential fields."""
    inventory = build_inventory()

    quota = analyse_object_quota(
        inventory,
        ObjectQuotaConfig(limit_override=100),
    )

    workbook = build_excel_workbook(
        capabilities=build_capabilities(),
        inventory=inventory,
        quota=quota,
    )

    headers: list[str] = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    headers.append(
                        str(cell.value).lower()
                    )

    serialised = " ".join(headers)

    assert "password" not in serialised
    assert "access_key" not in serialised
    assert "session_token" not in serialised


def test_existing_excel_report_is_protected(
    tmp_path: Path,
) -> None:
    """Existing workbooks should require explicit overwrite."""
    destination = tmp_path / "governance.xlsx"
    destination.write_bytes(b"existing")

    inventory = build_inventory()

    quota = analyse_object_quota(
        inventory,
        ObjectQuotaConfig(limit_override=100),
    )

    workbook = build_excel_workbook(
        capabilities=build_capabilities(),
        inventory=inventory,
        quota=quota,
    )

    with pytest.raises(ReportError):
        write_excel_report(
            workbook,
            destination,
        )


def test_excel_report_sanitises_illegal_control_characters(
    tmp_path: Path,
) -> None:
    """Excel-bound metadata should have illegal controls removed."""
    inventory = build_inventory(
        object_name="orders\x01",
    )

    quota = analyse_object_quota(
        inventory,
        ObjectQuotaConfig(limit_override=100),
    )

    workbook = build_excel_workbook(
        capabilities=build_capabilities(
            server_version="Redshift test\x00",
        ),
        inventory=inventory,
        quota=quota,
    )

    destination = tmp_path / "sanitised.xlsx"

    result = write_excel_report(
        workbook,
        destination,
    )

    reopened = load_workbook(
        result,
        read_only=True,
        data_only=True,
    )

    try:
        executive_summary = reopened["Executive Summary"]
        inventory_sheet = reopened["Object Inventory"]

        assert executive_summary["B7"].value == "Redshift test"
        assert inventory_sheet["C2"].value == "orders"
    finally:
        reopened.close()
